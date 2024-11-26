from aiogram import Router, F, Bot
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command
from datetime import datetime
from utils.keyboards.inline_builder import tickets_kb
from utils.keyboards.reply_kb import main_kb
from utils.requestsbd import create_spect_db, check_admin, get_all_users, get_sheets, get_suggestions
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup 
from aiogram.exceptions import TelegramBadRequest
from contextlib import suppress
from openpyxl import load_workbook

acom_router = Router()

class FormCreate(StatesGroup):
    name = State()
    desc = State()
    price = State()
    ticket_photo = State()
    place = State()
    date = State()

@acom_router.message(Command('apanel'))
async def apanel_cmd(msg: Message):
    if await check_admin(msg.from_user.id) == 1:
        kb = await tickets_kb()
        await msg.reply(f"<i>–°–ø–∏—Å–æ–∫ —Ç—ñ–∫–µ—Ç—ñ–≤ —Å—Ç–∞–Ω–æ–º –Ω–∞</i> <b>{datetime.now().strftime('%d-%m-%Y %H:%M:%S')}</b>", reply_markup=kb)


@acom_router.message(Command("zvit"))
async def cmd_zvit(msg: Message):
    result = await get_sheets()
    fn = "data/monday.xlsx"
    wb = load_workbook(fn)
    ws = wb['users']

    start_cell = 'A2'

    for index, res in enumerate(result):
        name_cell = ws.cell(row=int(start_cell[1:]) + index, column=1, value=res[0])  
        phone_cell = ws.cell(row=int(start_cell[1:]) + index, column=2, value=res[1])  
    wb.save(fn)

    await msg.reply_document(document = FSInputFile("data/monday.xlsx"))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None   
    wb.save(fn)
    wb.close()


@acom_router.message(Command('new_spect'))
async def create_spect(msg: Message, state: FSMContext):
    if await check_admin(msg.from_user.id) == 1:
        await msg.reply("–í–≤–µ–¥—ñ—Ç—å –Ω–∞–∑–≤—É –≤–∏—Å—Ç–∞–≤–∏:")
        await state.set_state(FormCreate.name)



@acom_router.message(FormCreate.name)
async def create_spect_name(msg: Message, state: FSMContext):
    await state.update_data(name = msg.text)
    await state.set_state(FormCreate.desc)
    await msg.reply("–¢–µ–ø–µ—Ä –≤–≤–µ–¥—ñ—Ç—å –æ–ø–∏—Å –≤–∏—Å—Ç–∞–≤–∏.")


@acom_router.message(FormCreate.desc)
async def create_spect_desc(msg: Message, state: FSMContext):
    await state.update_data(desc = msg.text)
    await state.set_state(FormCreate.price)
    await msg.reply("–¢–µ–ø–µ—Ä –≤–≤–µ–¥—ñ—Ç—å –≤–∞—Ä—Ç—ñ—Å—Ç—å.")

@acom_router.message(FormCreate.price)
async def create_spect_desc(msg: Message, state: FSMContext):
    await state.update_data(price = msg.text)
    await state.set_state(FormCreate.ticket_photo)
    await msg.reply("–¢–µ–ø–µ—Ä –≤–≤–µ–¥—ñ—Ç—å –ø–æ—Å–∏–ª–∞–Ω–Ω—è –Ω–∞ —Ñ–æ—Ç–æ –∫–≤–∏—Ç–∫–∞.")

@acom_router.message(FormCreate.ticket_photo)
async def create_spect_desc(msg: Message, state: FSMContext):
    await state.update_data(ticket_photo = msg.text)
    await state.set_state(FormCreate.place)
    await msg.reply("–¢–µ–ø–µ—Ä –≤–≤–µ–¥—ñ—Ç—å –º—ñ—Å—Ü–µ –ø—Ä–æ–≤–µ–¥–µ–Ω–Ω—è –≤–∏—Å—Ç–∞–≤–∏.")

@acom_router.message(FormCreate.place)
async def create_spect_place(msg: Message, state: FSMContext):
    await state.update_data(place = msg.text)
    await state.set_state(FormCreate.date)
    await msg.reply("–¢–µ–ø–µ—Ä –≤–≤–µ–¥—ñ—Ç—å –¥–∞—Ç—É –ø—Ä–æ–≤–µ–¥–µ–Ω–Ω—è –≤–∏—Å—Ç–∞–≤–∏. –£ —Ñ–æ—Ä–º–∞—Ç—ñ: dd-mm-yyyy H:M (01-02-2024 17:00)")

@acom_router.message(FormCreate.date)
async def create_spect_place(msg: Message, bot: Bot, state: FSMContext):
    data = await state.get_data()
    name = data['name']
    desc = data['desc']
    price = data['price']
    photo = data['ticket_photo']
    place = data['place']
    date = msg.text
    await state.clear()
    await create_spect_db(name, desc, place, date, price, photo)
    await msg.reply("–í–∏ —É—Å–ø—ñ—à–Ω–æ —Å—Ç–≤–æ—Ä–∏–ª–∏ –Ω–æ–≤—É –≤–∏—Å—Ç–∞–≤—É.", reply_markup = main_kb)
    users = await get_all_users()
    for user in users:
        with suppress(TelegramBadRequest):
            await bot.send_message(chat_id=user[1], text = f'<b>–ú–∏ –ø—ñ–¥–≥–æ—Ç—É–≤–∞–ª–∏ –Ω–æ–≤—É –≤–∏—Å—Ç–∞–≤—É:</b> <i>{name}</i>. <b>–•—É—Ç—á—ñ—à–µ –∑–∞—Ö–æ–¥—å—Ç–µ —É —Ä–æ–∑–¥—ñ–ª "–ù–∞–π–±–ª–∏–∂—á—ñ –≤–∏—Å—Ç–∞–≤–∏".</b>')


@acom_router.message(Command('suggest'))
async def asuggest(msg: Message):
    user_suggestions = await get_suggestions()
    user_list_text = "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ–π –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞:\n"
    for id, owner_id, owner_name, owner_phone, spect, date, in user_suggestions:
        user_info = f"üé´<b>‚Ññ</b>: {id} {owner_name}, <b>üì±–ù–æ–º–µ—Ä —Ç–µ–ª–µ—Ñ–æ–Ω—É:üì±</b> {owner_phone}, <b>üèõ–í–∏—Å—Ç–∞–≤–∞:</b> {spect}. üìÖ–î–∞—Ç–∞: {date}üé´\n\n‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî‚Äî-\n\n"
        user_list_text += user_info
    await msg.answer(user_list_text, parse_mode="html")