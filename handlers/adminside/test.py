from openpyxl import load_workbook
async def cmd_zvit():
fn = "E:\\lssp_bot\\handlers\\adminside\\monday.xlsx"
wb = load_workbook(fn)
ws = wb['data']

start_cell = 'A2'

for index, res in enumerate(result):
    name_cell = ws.cell(row=int(start_cell[1:]) + index, column=1, value=res[0])  # Запис імені
    phone_cell = ws.cell(row=int(start_cell[1:]) + index, column=2, value=res[1])  # Запис номера телефону

# Збереження змін та закриття книги
wb.save(fn)
wb.close()
